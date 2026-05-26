from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from tr_seo_contracts.module0 import Module0Request, SiteClassification, WebsiteProfile
from tr_seo_module0.business_context import BusinessContextAnalyzer, BusinessContextProfile


HEADER_ALIASES = {
    "keyword": {
        "keyword",
        "keywords",
        "phrase",
        "phrases",
        "searchterm",
        "searchterms",
        "query",
        "queries",
    },
    "search_volume": {
        "searchvolume",
        "volume",
        "avgmonthlysearches",
        "averagemonthlysearches",
        "monthlysearches",
        "sv",
    },
    "keyword_difficulty": {
        "kd",
        "difficulty",
        "keyworddifficulty",
        "seodifficulty",
    },
    "cpc": {
        "cpc",
        "avgcpc",
        "costperclick",
        "averagecpc",
    },
    "current_position": {
        "position",
        "pos",
        "currentposition",
        "rank",
        "currentrank",
    },
    "mapped_url": {
        "url",
        "mappedurl",
        "currenturl",
        "currentpageurl",
        "existingurl",
        "targeturl",
        "rankingurl",
        "pageurl",
        "page",
        "landingpage",
        "landingpageurl",
        "destinationurl",
    },
}

PRICING_TOKENS = {"cost", "price", "pricing", "quote", "quotes", "estimate", "estimates"}
QUESTION_TOKENS = {"how", "what", "why", "when", "who", "where", "does", "do", "can", "should", "is"}
LOCAL_TOKENS = {"near", "me", "melbourne", "sydney", "brisbane", "perth", "adelaide"}
PROBLEM_TOKENS = {"emergency", "leaking", "broken", "rusted", "damaged", "overflowing", "sagging"}


@dataclass(slots=True)
class UploadedKeywordRow:
    keyword: str
    search_volume: int
    keyword_difficulty: int
    cpc: float | None = None
    current_position: int | None = None
    mapped_url: str | None = None
    original_row: dict[str, str] = field(default_factory=dict)


@dataclass(slots=True)
class KeywordRejection:
    keyword: str
    reason: str
    details: str = ""


@dataclass(slots=True)
class AcceptedKeywordRow:
    keyword: str
    search_volume: int
    keyword_difficulty: int
    cpc: float | None = None
    current_position: int | None = None
    mapped_url: str | None = None
    confidence_score: float = 0.0
    quality_score: float = 0.0
    notes: list[str] = field(default_factory=list)


@dataclass(slots=True)
class ManualKeywordUploadResult:
    parser_used: str
    file_name: str
    raw_rows: int
    accepted_keywords: list[AcceptedKeywordRow]
    rejected_keywords: list[KeywordRejection]
    notes: list[str] = field(default_factory=list)


class ManualKeywordUploadProcessor:
    SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".csv"}

    def __init__(self, context_analyzer: BusinessContextAnalyzer | None = None) -> None:
        self.context_analyzer = context_analyzer or BusinessContextAnalyzer()

    def process(
        self,
        *,
        filename: str,
        content: bytes,
        request: Module0Request,
        site_classification: SiteClassification | None,
        website_profile: WebsiteProfile | None,
        keyword_limit: int,
    ) -> ManualKeywordUploadResult:
        extension = Path(filename).suffix.lower()
        if extension not in self.SUPPORTED_EXTENSIONS:
            raise ValueError(f"Unsupported keyword upload format: {extension}")

        parsed_rows, parser_used, notes = self._parse_file(extension=extension, content=content)
        context = self.context_analyzer.build(
            request=request,
            site_classification=site_classification,
            website_profile=website_profile,
        )
        keep_product_terms = context.business_type.value in {"ecommerce", "hybrid"}
        accepted: list[AcceptedKeywordRow] = []
        rejected: list[KeywordRejection] = []
        seen_exact: set[str] = set()
        seen_signatures: set[str] = set()

        for row in parsed_rows:
            rejection = self._validate_row(
                row=row,
                context=context,
                keep_product_terms=keep_product_terms,
                seen_exact=seen_exact,
                seen_signatures=seen_signatures,
            )
            if rejection is not None:
                rejected.append(rejection)
                continue

            normalized_keyword = self.context_analyzer.normalize_query(row.keyword, context)
            naturalness = self.context_analyzer.naturalness_score(normalized_keyword, context)
            relevance = self.context_analyzer.business_relevance_score(
                normalized_keyword,
                context,
                keep_product_terms=keep_product_terms,
            )
            confidence = self._confidence_from_metrics(row, naturalness, relevance)
            quality = round((naturalness * 0.45) + (relevance * 0.55), 2)
            seen_exact.add(normalized_keyword)
            seen_signatures.add(self._semantic_signature(normalized_keyword, context))
            accepted.append(
                AcceptedKeywordRow(
                    keyword=normalized_keyword,
                    search_volume=row.search_volume,
                    keyword_difficulty=row.keyword_difficulty,
                    current_position=row.current_position,
                    cpc=row.cpc,
                    mapped_url=row.mapped_url,
                    confidence_score=confidence,
                    quality_score=quality,
                    notes=[
                        "Imported from manual SEMrush keyword upload.",
                        f"Naturalness score: {naturalness:.2f}.",
                        f"Business relevance score: {relevance:.2f}.",
                    ],
                )
            )

        accepted.sort(
            key=lambda item: (
                -item.search_volume,
                -item.quality_score,
                item.current_position or 999,
                item.keyword,
            )
        )
        limited_accepted = accepted[:keyword_limit]
        if len(accepted) > keyword_limit:
            notes.append(
                f"Accepted {len(accepted)} relevant keywords from the upload, then limited the working set to {keyword_limit}."
            )
        notes.append(
            f"Filtered {len(parsed_rows)} uploaded keyword rows into {len(limited_accepted)} accepted and {len(rejected)} rejected rows."
        )
        return ManualKeywordUploadResult(
            parser_used=parser_used,
            file_name=filename,
            raw_rows=len(parsed_rows),
            accepted_keywords=limited_accepted,
            rejected_keywords=rejected,
            notes=notes,
        )

    def _parse_file(
        self,
        *,
        extension: str,
        content: bytes,
    ) -> tuple[list[UploadedKeywordRow], str, list[str]]:
        if extension in {".xlsx", ".xls"}:
            return self._parse_spreadsheet(content)
        return self._parse_csv(content)

    def _parse_spreadsheet(self, content: bytes) -> tuple[list[UploadedKeywordRow], str, list[str]]:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        parsed_rows: list[UploadedKeywordRow] = []
        notes: list[str] = []
        for sheet in workbook.worksheets:
            rows = [
                ["" if cell is None else str(cell).strip() for cell in row]
                for row in sheet.iter_rows(values_only=True)
            ]
            sheet_rows = self._rows_from_matrix(rows)
            if sheet_rows:
                parsed_rows.extend(sheet_rows)
            else:
                notes.append(f"No recognizable keyword table found in sheet '{sheet.title}'.")
        return parsed_rows, "openpyxl", notes

    def _parse_csv(self, content: bytes) -> tuple[list[UploadedKeywordRow], str, list[str]]:
        text = content.decode("utf-8", errors="ignore")
        sample = text[:4000]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(io.StringIO(text), dialect)
        rows = [["" if cell is None else str(cell).strip() for cell in row] for row in reader]
        return self._rows_from_matrix(rows), "csv", []

    def _rows_from_matrix(self, rows: list[list[str]]) -> list[UploadedKeywordRow]:
        if not rows:
            return []
        header_index, header_map = self._detect_header(rows[:10])
        if header_index is None or "keyword" not in header_map:
            return []

        parsed: list[UploadedKeywordRow] = []
        for row in rows[header_index + 1 :]:
            if not any(cell.strip() for cell in row):
                continue
            keyword = self._cell(row, header_map.get("keyword"))
            if not keyword:
                continue
            parsed.append(
                UploadedKeywordRow(
                    keyword=keyword,
                    search_volume=self._to_int(self._cell(row, header_map.get("search_volume"))),
                    keyword_difficulty=min(100, self._to_int(self._cell(row, header_map.get("keyword_difficulty")))),
                    cpc=self._to_optional_float(self._cell(row, header_map.get("cpc"))),
                    current_position=self._to_optional_int(self._cell(row, header_map.get("current_position"))),
                    mapped_url=self._normalize_url_value(
                        self._cell(row, header_map.get("mapped_url")) or self._infer_url_from_row(row)
                    ),
                    original_row={
                        f"column_{index}": value
                        for index, value in enumerate(row)
                        if value not in {"", None}
                    },
                )
            )
        return parsed

    def _detect_header(self, rows: list[list[str]]) -> tuple[int | None, dict[str, int]]:
        best_index: int | None = None
        best_mapping: dict[str, int] = {}
        best_score = 0

        for index, row in enumerate(rows):
            mapping: dict[str, int] = {}
            score = 0
            for column_index, value in enumerate(row):
                normalized = self._normalize_header(value)
                if not normalized:
                    continue
                for field_name, aliases in HEADER_ALIASES.items():
                    if normalized in aliases and field_name not in mapping:
                        mapping[field_name] = column_index
                        score += 3 if field_name == "keyword" else 1
            if "keyword" in mapping and score > best_score:
                best_index = index
                best_mapping = mapping
                best_score = score

        return best_index, best_mapping

    def _validate_row(
        self,
        *,
        row: UploadedKeywordRow,
        context: BusinessContextProfile,
        keep_product_terms: bool,
        seen_exact: set[str],
        seen_signatures: set[str],
    ) -> KeywordRejection | None:
        normalized_keyword = self.context_analyzer.normalize_query(row.keyword, context)
        if not normalized_keyword:
            return KeywordRejection(keyword=row.keyword, reason="empty_keyword", details="Keyword cell was empty after normalization.")
        if normalized_keyword in seen_exact:
            return KeywordRejection(keyword=row.keyword, reason="exact_duplicate", details="Duplicate normalized keyword.")

        semantic_signature = self._semantic_signature(normalized_keyword, context)
        if semantic_signature in seen_signatures:
            return KeywordRejection(keyword=row.keyword, reason="near_duplicate", details="Near-duplicate semantic pattern.")

        if not self.context_analyzer.is_valid_query(
            normalized_keyword,
            context,
            keep_product_terms=keep_product_terms,
        ):
            naturalness = self.context_analyzer.naturalness_score(normalized_keyword, context)
            relevance = self.context_analyzer.business_relevance_score(
                normalized_keyword,
                context,
                keep_product_terms=keep_product_terms,
            )
            if naturalness < 0.62:
                return KeywordRejection(
                    keyword=row.keyword,
                    reason="unnatural_query",
                    details=f"Naturalness score {naturalness:.2f} was below the required threshold.",
                )
            return KeywordRejection(
                keyword=row.keyword,
                reason="not_business_aligned",
                details=f"Business relevance score {relevance:.2f} was below the required threshold.",
            )

        quality = (
            self.context_analyzer.naturalness_score(normalized_keyword, context) * 0.45
            + self.context_analyzer.business_relevance_score(
                normalized_keyword,
                context,
                keep_product_terms=keep_product_terms,
            )
            * 0.55
        )
        if quality < 0.66:
            return KeywordRejection(
                keyword=row.keyword,
                reason="below_quality_threshold",
                details=f"Combined quality score {quality:.2f} was below threshold.",
            )
        return None

    def _semantic_signature(self, value: str, context: BusinessContextProfile) -> str:
        synonym_map = {
            "pricing": "pricing",
            "price": "pricing",
            "cost": "pricing",
            "quote": "pricing",
            "quotes": "pricing",
            "estimate": "pricing",
            "estimates": "pricing",
            "repair": "repair",
            "repairs": "repair",
            "replacement": "replacement",
            "replacements": "replacement",
            "install": "installation",
            "installation": "installation",
            "installations": "installation",
            "near": "local",
            "me": "local",
        }
        tokens = [
            synonym_map.get(token, token)
            for token in value.split()
            if token not in context.brand_tokens
        ]
        normalized_tokens: list[str] = []
        for token in tokens:
            if token in QUESTION_TOKENS:
                normalized_tokens.append("question")
                continue
            if token in PRICING_TOKENS:
                normalized_tokens.append("pricing")
                continue
            if token in LOCAL_TOKENS or token in context.location_tokens:
                normalized_tokens.append(token)
                continue
            if token in PROBLEM_TOKENS:
                normalized_tokens.append(token)
                continue
            normalized_tokens.append(token)
        compacted = list(dict.fromkeys(normalized_tokens))
        return " ".join(compacted)

    def _confidence_from_metrics(
        self,
        row: UploadedKeywordRow,
        naturalness: float,
        relevance: float,
    ) -> float:
        confidence = 0.72 + min(0.12, (naturalness + relevance) * 0.08)
        if row.search_volume > 0:
            confidence += 0.04
        if row.keyword_difficulty > 0:
            confidence += 0.03
        if row.cpc is not None:
            confidence += 0.03
        return round(min(confidence, 0.97), 2)

    def _normalize_header(self, value: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())

    def _normalize_url_value(self, value: str) -> str | None:
        normalized = str(value or "").strip()
        return normalized or None

    def _infer_url_from_row(self, row: list[str]) -> str | None:
        for value in row:
            text = str(value or "").strip()
            if not text:
                continue
            lowered = text.lower()
            if lowered.startswith(("http://", "https://", "/")):
                return text
            if re.search(r"\.(com|com\.au|au|net|org)(/|$)", lowered):
                return text
        return None

    def _cell(self, row: list[str], index: int | None) -> str:
        if index is None or index >= len(row):
            return ""
        return str(row[index] or "").strip()

    def _to_int(self, value: str | None) -> int:
        cleaned = re.sub(r"[^0-9]", "", str(value or ""))
        if not cleaned:
            return 0
        return int(cleaned)

    def _to_optional_int(self, value: str | None) -> int | None:
        cleaned = re.sub(r"[^0-9]", "", str(value or ""))
        if not cleaned:
            return None
        return int(cleaned)

    def _to_optional_float(self, value: str | None) -> float | None:
        text = str(value or "").strip()
        if not text:
            return None
        text = text.replace(",", "")
        match = re.search(r"\d+(?:\.\d+)?", text)
        if not match:
            return None
        return float(match.group(0))
