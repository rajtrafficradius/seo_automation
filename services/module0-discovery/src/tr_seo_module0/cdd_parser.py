from __future__ import annotations

import csv
import io
import re
from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

from tr_seo_contracts.module0 import CDDExtraction, CDDFileMeta


class CDDParser:
    SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".xls", ".csv"}
    DOMAIN_PATTERN = re.compile(
        r"(?:https?://)?(?:www\.)?([a-z0-9][a-z0-9.-]+\.[a-z]{2,})(?:/[^\s|]*)?",
        re.IGNORECASE,
    )

    def parse(self, file_meta: CDDFileMeta, content: bytes) -> CDDExtraction:
        extension = file_meta.extension.lower()
        if extension not in self.SUPPORTED_EXTENSIONS:
            return CDDExtraction(
                text_preview="",
                parser_used="unsupported",
                warnings=[f"Unsupported file type: {extension}"],
            )

        if extension == ".pdf":
            return self._parse_pdf(content)
        if extension == ".docx":
            return self._parse_docx(content)
        if extension in {".xlsx", ".xls"}:
            return self._parse_spreadsheet(content)
        return self._parse_csv(content)

    def _parse_pdf(self, content: bytes) -> CDDExtraction:
        reader = PdfReader(io.BytesIO(content))
        extracted = "\n".join((page.extract_text() or "").strip() for page in reader.pages).strip()
        return self._build_extraction(extracted, "pypdf")

    def _parse_docx(self, content: bytes) -> CDDExtraction:
        document = Document(io.BytesIO(content))
        extracted = "\n".join(paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text)
        return self._build_extraction(extracted, "python-docx")

    def _parse_spreadsheet(self, content: bytes) -> CDDExtraction:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        rows: list[str] = []
        for sheet in workbook.worksheets:
            rows.append(f"[Sheet] {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [str(cell).strip() for cell in row if cell not in (None, "")]
                if values:
                    rows.append(" | ".join(values))
        return self._build_extraction("\n".join(rows), "openpyxl")

    def _parse_csv(self, content: bytes) -> CDDExtraction:
        text = content.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        rows = [" | ".join(cell.strip() for cell in row if cell.strip()) for row in reader]
        extracted = "\n".join(row for row in rows if row)
        return self._build_extraction(extracted, "csv")

    def _build_extraction(self, extracted_text: str, parser_used: str) -> CDDExtraction:
        preview = extracted_text[:1200]
        sections = self._detect_sections(extracted_text)
        detected_domains = self._extract_domains(extracted_text)
        competitor_hints = self._extract_competitor_hints(extracted_text)
        warnings: list[str] = []
        if not extracted_text.strip():
            warnings.append("No extractable text was found in the uploaded CDD file.")
        return CDDExtraction(
            text_preview=preview,
            sections_detected=sections,
            parser_used=parser_used,
            detected_domains=detected_domains,
            competitor_hints=competitor_hints,
            warnings=warnings,
        )

    def _detect_sections(self, extracted_text: str) -> list[str]:
        lowered = extracted_text.lower()
        known_sections = [
            "business overview",
            "services",
            "products",
            "target locations",
            "goals",
            "competitors",
            "brand",
            "notes",
        ]
        return [section for section in known_sections if section in lowered]

    def _extract_domains(self, extracted_text: str) -> list[str]:
        domains: list[str] = []
        seen: set[str] = set()
        for match in self.DOMAIN_PATTERN.finditer(extracted_text):
            domain = match.group(1).lower().strip(".")
            if domain.startswith("www."):
                domain = domain[4:]
            if domain not in seen:
                seen.add(domain)
                domains.append(domain)
        return domains[:50]

    def _extract_competitor_hints(self, extracted_text: str) -> list[str]:
        hints: list[str] = []
        seen: set[str] = set()
        lines = [line.strip() for line in extracted_text.splitlines() if line.strip()]

        for index, line in enumerate(lines):
            lowered = line.lower()
            if "competitor" not in lowered:
                continue

            window = [line]
            window.extend(lines[index + 1 : index + 6])
            for raw in window:
                candidates = self._split_competitor_line(raw)
                for candidate in candidates:
                    normalized = candidate.strip()
                    normalized_lower = normalized.lower()
                    if (
                        not normalized
                        or "competitor" in normalized_lower
                        or normalized_lower in seen
                    ):
                        continue
                    seen.add(normalized_lower)
                    hints.append(normalized)

        for domain in self._extract_domains(extracted_text):
            if domain not in seen:
                seen.add(domain)
                hints.append(domain)

        return hints[:40]

    def _split_competitor_line(self, line: str) -> list[str]:
        chunks = re.split(r"\s*\|\s*|\n|,|;", line)
        values: list[str] = []
        for chunk in chunks:
            candidate = chunk.strip().strip("*").strip()
            if not candidate:
                continue
            values.append(candidate)
        return values

    def make_file_meta(self, filename: str, content_type: str, size_bytes: int) -> CDDFileMeta:
        extension = Path(filename).suffix.lower()
        return CDDFileMeta(
            filename=filename,
            content_type=content_type or "application/octet-stream",
            size_bytes=size_bytes,
            extension=extension,
        )
