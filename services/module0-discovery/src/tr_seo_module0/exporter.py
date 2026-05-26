from __future__ import annotations

from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from tr_seo_contracts.module0 import Module0ExportFile, Module0Exports, Module0Response


class Module0Exporter:
    HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    LABEL_FONT = Font(bold=True)

    def export(self, response: Module0Response, output_dir: Path, run_id: str) -> Module0Exports:
        output_dir.mkdir(parents=True, exist_ok=True)
        company_prefix = self._company_prefix(response.request.brand_name)

        export_files = {
            "cdd_extraction": output_dir / f"{company_prefix}_cdd_extraction.xlsx",
            "site_classification": output_dir / f"{company_prefix}_site_classification.xlsx",
            "website_profile": output_dir / f"{company_prefix}_website_profile.xlsx",
            "semrush_snapshot": output_dir / f"{company_prefix}_semrush_snapshot.xlsx",
            "competitive_intelligence": output_dir / f"{company_prefix}_competitive_intelligence.xlsx",
            "keyword_universe": output_dir / f"{company_prefix}_master_keyword_universe.xlsx",
            "keyword_clusters": output_dir / f"{company_prefix}_keyword_clusters.xlsx",
            "quick_wins": output_dir / f"{company_prefix}_quick_wins.xlsx",
            "tam_dataset": output_dir / f"{company_prefix}_tam_dataset.xlsx",
            "url_architecture_map": output_dir / f"{company_prefix}_url_architecture_map.xlsx",
            "minimum_effort_points": output_dir / f"{company_prefix}_minimum_effort_points.xlsx",
            "ai_sov_baseline": output_dir / f"{company_prefix}_ai_sov_baseline.xlsx",
            "fan_out_map": output_dir / f"{company_prefix}_fan_out_map.xlsx",
            "entity_authority_baseline": output_dir / f"{company_prefix}_entity_authority_baseline.xlsx",
            "warnings_errors": output_dir / f"{company_prefix}_warnings_errors.xlsx",
            "full_run_workbook": output_dir / f"{company_prefix}_module0_full_export.xlsx",
        }

        self._write_cdd_extraction(response, export_files["cdd_extraction"])
        self._write_site_classification(response, export_files["site_classification"])
        self._write_website_profile(response, export_files["website_profile"])
        self._write_semrush_snapshot(response, export_files["semrush_snapshot"])
        self._write_competitive_intelligence(response, export_files["competitive_intelligence"])
        self._write_keyword_universe(response, export_files["keyword_universe"])
        self._write_keyword_clusters(response, export_files["keyword_clusters"])
        self._write_quick_wins(response, export_files["quick_wins"])
        self._write_tam_dataset(response, export_files["tam_dataset"])
        self._write_url_architecture_map(response, export_files["url_architecture_map"])
        self._write_minimum_effort_points(response, export_files["minimum_effort_points"])
        self._write_ai_sov_baseline(response, export_files["ai_sov_baseline"])
        self._write_fan_out_map(response, export_files["fan_out_map"])
        self._write_entity_authority_baseline(response, export_files["entity_authority_baseline"])
        self._write_warnings_errors(response, export_files["warnings_errors"])
        self._write_full_workbook(response, export_files["full_run_workbook"])

        return Module0Exports(
            cdd_extraction=self._export_file(export_files["cdd_extraction"]),
            site_classification=self._export_file(export_files["site_classification"]),
            website_profile=self._export_file(export_files["website_profile"]),
            semrush_snapshot=self._export_file(export_files["semrush_snapshot"]),
            competitive_intelligence=self._export_file(export_files["competitive_intelligence"]),
            keyword_universe=self._export_file(export_files["keyword_universe"]),
            keyword_clusters=self._export_file(export_files["keyword_clusters"]),
            quick_wins=self._export_file(export_files["quick_wins"]),
            tam_dataset=self._export_file(export_files["tam_dataset"]),
            url_architecture_map=self._export_file(export_files["url_architecture_map"]),
            minimum_effort_points=self._export_file(export_files["minimum_effort_points"]),
            ai_sov_baseline=self._export_file(export_files["ai_sov_baseline"]),
            fan_out_map=self._export_file(export_files["fan_out_map"]),
            entity_authority_baseline=self._export_file(export_files["entity_authority_baseline"]),
            warnings_errors=self._export_file(export_files["warnings_errors"]),
            full_run_workbook=self._export_file(export_files["full_run_workbook"]),
        )

    def _company_prefix(self, brand_name: str) -> str:
        cleaned = re.sub(r"[^A-Za-z0-9]+", "_", brand_name.strip().lower())
        cleaned = re.sub(r"_+", "_", cleaned).strip("_")
        return cleaned or "module0_export"

    def _export_file(self, target: Path) -> Module0ExportFile:
        return Module0ExportFile(
            filename=target.name,
            download_url=f"/api/v1/module0/downloads/{target.name}",
        )

    def _write_cdd_extraction(self, response: Module0Response, target: Path) -> None:
        workbook = Workbook()
        overview = workbook.active
        overview.title = "CDD Extraction"
        self._append_key_value_rows(
            overview,
            [
                ("Filename", response.cdd_file.filename),
                ("Content Type", response.cdd_file.content_type),
                ("Extension", response.cdd_file.extension),
                ("Size Bytes", response.cdd_file.size_bytes),
                ("Parser Used", response.cdd_extraction.parser_used),
                ("Sections Detected", ", ".join(response.cdd_extraction.sections_detected)),
            ],
        )
        self._add_single_column_sheet(
            workbook,
            "Text Preview",
            "Text Preview",
            [response.cdd_extraction.text_preview],
            width=120,
        )
        self._add_single_column_sheet(
            workbook,
            "Warnings",
            "Warning",
            response.cdd_extraction.warnings or ["No parser warnings."],
            width=100,
        )
        workbook.save(target)

    def _write_site_classification(self, response: Module0Response, target: Path) -> None:
        workbook = Workbook()
        overview = workbook.active
        overview.title = "Site Classification"
        classification = response.site_classification
        self._append_key_value_rows(
            overview,
            [
                ("Detected Domain", classification.detected_domain),
                ("Business Type", classification.business_type.value),
                ("Industry Category", classification.industry_category),
                ("Geographic Target", classification.geographic_target),
                ("Language", classification.language),
                ("CMS", classification.cms),
                ("CMS Version", classification.cms_version or ""),
                ("Site Scale Tier", classification.site_scale_tier),
                ("Page Builder", classification.page_builder or ""),
                ("Sitemap URL", classification.sitemap_url or ""),
                ("Theme / Template", classification.theme_or_template or ""),
                ("Confidence Score", classification.confidence_score),
            ],
        )
        self._add_single_column_sheet(
            workbook,
            "Active Components",
            "Component",
            classification.active_components or ["No active components detected."],
            width=80,
        )
        self._add_single_column_sheet(
            workbook,
            "Notes",
            "Note",
            classification.notes or ["No notes."],
            width=100,
        )
        workbook.save(target)

    def _write_website_profile(self, response: Module0Response, target: Path) -> None:
        workbook = Workbook()
        overview = workbook.active
        overview.title = "Website Profile"
        profile = response.website_profile
        self._append_key_value_rows(
            overview,
            [
                ("Homepage Status", profile.homepage_status_code or ""),
                ("Final Status", profile.final_status_code or ""),
                ("Response Time (ms)", profile.response_time_ms or ""),
                ("Redirect Count", profile.redirect_count),
                ("Homepage Title", profile.homepage_title or ""),
                ("Meta Description", profile.meta_description or ""),
                ("Canonical URL", profile.canonical_url or ""),
                ("Word Count", profile.word_count),
                ("H1 Count", profile.h1_count),
                ("Primary Schema Type", profile.primary_schema_type or ""),
                ("Mobile Friendly", profile.mobile_friendly),
                ("Indexable", profile.indexable),
                ("Robots URL", profile.robots_txt.url),
                ("Robots Fetched", "Yes" if profile.robots_txt.fetched else "No"),
                ("Robots Status", profile.robots_txt.status_code or ""),
                ("Allows AI Crawlers", profile.robots_txt.allows_ai_crawlers),
                ("Sitemap Discovered", "Yes" if profile.sitemap.discovered else "No"),
                ("Sitemap Count", profile.sitemap.fetched_count),
                ("Sitemap URL Count", profile.sitemap.url_count),
                ("Total URLs", profile.url_inventory.total_urls),
            ],
        )
        self._add_key_value_sheet(
            workbook,
            "Headers",
            [(key, value) for key, value in profile.important_headers.items()] or [("Info", "No headers captured.")],
        )
        self._add_single_column_sheet(
            workbook,
            "Schema Types",
            "Schema Type",
            profile.detected_schema_types or ["No schema types detected."],
        )
        self._add_single_column_sheet(
            workbook,
            "Social Profiles",
            "Profile URL",
            profile.social_profile_links or ["No social links detected."],
            width=100,
        )
        self._add_single_column_sheet(
            workbook,
            "Broken Internal Links",
            "URL",
            profile.broken_internal_links or ["No broken internal links found in the safe crawl sample."],
            width=100,
        )
        self._add_single_column_sheet(
            workbook,
            "Robots Notes",
            "Note",
            profile.robots_txt.notes or ["No robots notes."],
            width=100,
        )
        self._add_single_column_sheet(
            workbook,
            "Sitemap URLs",
            "Sitemap URL",
            profile.sitemap.sitemap_urls or ["No sitemap URLs detected."],
            width=100,
        )
        self._add_single_column_sheet(
            workbook,
            "Sitemap Sample URLs",
            "Sample URL",
            profile.sitemap.sample_urls or ["No sample URLs captured."],
            width=100,
        )
        inventory_sheet = workbook.create_sheet("URL Inventory")
        self._write_headers(inventory_sheet, 1, ["Section", "URL Count"])
        for item in profile.url_inventory.top_sections:
            inventory_sheet.append([item.section, item.url_count])
        self._set_column_widths(inventory_sheet, [32, 16])
        self._add_single_column_sheet(
            workbook,
            "Service-like URLs",
            "URL",
            profile.url_inventory.service_like_urls or ["No service-like URLs detected."],
            width=100,
        )
        self._add_single_column_sheet(
            workbook,
            "Location-like URLs",
            "URL",
            profile.url_inventory.location_like_urls or ["No location-like URLs detected."],
            width=100,
        )
        workbook.save(target)

    def _write_semrush_snapshot(self, response: Module0Response, target: Path) -> None:
        workbook = Workbook()
        overview = workbook.active
        overview.title = "SEMrush Snapshot"
        semrush = response.semrush
        self._append_key_value_rows(
            overview,
            [
                ("Configured", semrush.configured),
                ("Region Database", semrush.region_database),
                ("Status", semrush.status),
                ("Data Source", semrush.data_source),
                ("Fallback Used", semrush.fallback_used),
                ("Is Estimated", semrush.is_estimated),
                ("Warning Message", semrush.warning_message or ""),
                ("Keyword Limit", semrush.keyword_limit or ""),
                ("Estimated Traffic", semrush.estimated_monthly_traffic or ""),
                ("Organic Keyword Count", semrush.organic_keyword_count or ""),
                ("Competitors Evaluated", semrush.competitors_evaluated),
                ("Source File", semrush.source_file_name or ""),
                ("Source File Extension", semrush.source_file_extension or ""),
                ("Raw Keyword Rows", semrush.raw_keyword_rows or ""),
                ("Accepted Keyword Rows", semrush.accepted_keyword_rows or ""),
                ("Rejected Keyword Rows", semrush.rejected_keyword_rows or ""),
            ],
        )
        self._add_single_column_sheet(
            workbook,
            "Notes",
            "Note",
            semrush.notes or ["No SEMrush notes."],
            width=100,
        )
        self._add_single_column_sheet(
            workbook,
            "Rejected Keyword Examples",
            "Rejected Keyword",
            semrush.rejected_keyword_examples or ["No rejected keyword examples recorded."],
            width=100,
        )
        history_sheet = workbook.create_sheet("Traffic History")
        self._write_headers(history_sheet, 1, ["Month Index", "Estimated Traffic"])
        for index, value in enumerate(semrush.estimated_monthly_traffic_history, start=1):
            history_sheet.append([index, value])
        self._set_column_widths(history_sheet, [16, 20])
        workbook.save(target)

    def _write_competitive_intelligence(self, response: Module0Response, target: Path) -> None:
        workbook = Workbook()
        competitors = workbook.active
        competitors.title = "Competitors"
        self._write_headers(
            competitors,
            1,
            [
                "Domain",
                "Name",
                "Source",
                "Reason for Selection",
                "Likely Services",
                "Content Gaps",
                "Service Gaps",
                "Estimated Strength",
                "Confidence",
                "Is Estimated",
                "Anomaly Filtered",
                "Competition Level",
                "Shared Keywords",
                "Keyword Sample",
                "Notes",
            ],
        )
        for item in response.competitive_intelligence.top_competitors:
            competitors.append(
                [
                    item.domain,
                    item.name,
                    item.source,
                    item.reason_for_selection,
                    ", ".join(item.likely_services),
                    ", ".join(item.content_gaps),
                    ", ".join(item.service_gaps),
                    item.estimated_strength,
                    item.confidence_score,
                    "Yes" if item.is_estimated else "No",
                    "Yes" if item.anomaly_filtered else "No",
                    item.competition_level or "",
                    item.shared_keywords or "",
                    ", ".join(item.keyword_sample),
                    " | ".join(item.notes),
                ]
            )
        self._set_column_widths(competitors, [32, 24, 18, 40, 32, 32, 32, 16, 14, 14, 16, 16, 16, 46, 48])

        self._add_gap_sheet(workbook, "Service Gaps", response.competitive_intelligence.service_gaps)
        self._add_gap_sheet(workbook, "Content Gaps", response.competitive_intelligence.content_gaps)
        local_pages = workbook.create_sheet("Local Page Competitors")
        self._write_headers(local_pages, 1, ["Page", "Competitor Domain"])
        for page, domains in response.competitive_intelligence.local_page_competitors.items():
            if domains:
                for domain in domains:
                    local_pages.append([page, domain])
            else:
                local_pages.append([page, ""])
        self._set_column_widths(local_pages, [48, 34])
        self._add_single_column_sheet(
            workbook,
            "Filtered Domains",
            "Domain",
            response.competitive_intelligence.filtered_domains or ["No filtered domains."],
            width=60,
        )
        workbook.save(target)

    def _add_gap_sheet(self, workbook: Workbook, title: str, gaps) -> None:
        sheet = workbook.create_sheet(title)
        self._write_headers(
            sheet,
            1,
            ["Gap Type", "Label", "Supporting Keywords", "Rationale", "Opportunity Score"],
        )
        if gaps:
            for gap in gaps:
                sheet.append(
                    [
                        gap.gap_type,
                        gap.label,
                        ", ".join(gap.supporting_keywords),
                        gap.rationale,
                        gap.opportunity_score,
                    ]
                )
        else:
            sheet.append(["", "No gaps identified.", "", "", ""])
        self._set_column_widths(sheet, [18, 28, 48, 64, 18])

    def _write_keyword_universe(self, response: Module0Response, target: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Keyword Universe"
        self._append_key_value_rows(
            sheet,
            [
                ("Brand Name", response.request.brand_name),
                ("Website URL", str(response.request.website_url)),
                ("Target Country", response.request.target_country),
                ("SEMrush Source", response.semrush.data_source),
                ("SEMrush Status", response.semrush.status),
                ("Keyword Limit", response.semrush.keyword_limit or ""),
                ("Fallback Warning", response.semrush.warning_message or ""),
                ("TAM Estimate", response.tam_dataset.total_monthly_search_volume),
            ],
        )
        start_row = 10
        headers = [
            "Keyword",
            "Cluster ID",
            "Intent",
            "Priority",
            "Search Volume",
            "Keyword Difficulty",
            "CPC",
            "Current Position",
            "Source",
            "Mapped URL",
            "Quick Win",
            "AI Trigger Rate",
            "Confidence Score",
            "Quality Score",
            "Is Estimated",
        ]
        self._write_headers(sheet, start_row, headers)
        for keyword in response.master_keyword_universe:
            sheet.append(
                [
                    keyword.keyword,
                    keyword.cluster_id or "",
                    keyword.intent.value,
                    keyword.priority.value,
                    keyword.search_volume,
                    keyword.keyword_difficulty,
                    keyword.cpc or "",
                    keyword.current_position or "",
                    keyword.source,
                    keyword.mapped_url or "",
                    "Yes" if keyword.quick_win else "No",
                    keyword.ai_answer_trigger_rate or "",
                    keyword.confidence_score,
                    keyword.quality_score,
                    "Yes" if keyword.is_estimated else "No",
                ]
            )
        sheet.freeze_panes = f"A{start_row + 1}"
        sheet.auto_filter.ref = f"A{start_row}:O{sheet.max_row}"
        self._set_column_widths(sheet, [34, 20, 18, 12, 16, 20, 10, 18, 28, 34, 12, 14, 16, 16, 14])
        workbook.save(target)

    def _write_keyword_clusters(self, response: Module0Response, target: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Keyword Clusters"
        self._write_headers(
            sheet,
            1,
            ["Cluster ID", "Label", "Intent", "Primary Keyword", "Keywords", "Total Search Volume", "Suggested URL"],
        )
        for cluster in response.keyword_clusters:
            sheet.append(
                [
                    cluster.cluster_id,
                    cluster.label,
                    cluster.intent.value,
                    cluster.primary_keyword,
                    ", ".join(cluster.keywords),
                    cluster.total_search_volume,
                    cluster.suggested_url or "",
                ]
            )
        self._set_column_widths(sheet, [24, 28, 18, 32, 60, 20, 40])
        workbook.save(target)

    def _write_quick_wins(self, response: Module0Response, target: Path) -> None:
        workbook = Workbook()
        overview = workbook.active
        overview.title = "Quick Wins"
        self._append_key_value_rows(
            overview,
            [
                ("Total Quick Wins", response.quick_wins.total_count),
                ("SEMrush Data Source", response.semrush.data_source),
            ],
        )
        details = workbook.create_sheet("Keywords")
        self._write_headers(
            details,
            1,
            ["Keyword", "Priority", "Intent", "Search Volume", "Keyword Difficulty", "Current Position", "Mapped URL", "Source"],
        )
        for item in response.quick_wins.keywords:
            details.append(
                [
                    item.keyword,
                    item.priority.value,
                    item.intent.value,
                    item.search_volume,
                    item.keyword_difficulty,
                    item.current_position or "",
                    item.mapped_url or "",
                    item.source,
                ]
            )
        self._set_column_widths(details, [36, 12, 18, 16, 20, 18, 36, 24])
        workbook.save(target)

    def _write_tam_dataset(self, response: Module0Response, target: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "TAM Dataset"
        tam = response.tam_dataset
        self._append_key_value_rows(
            sheet,
            [
                ("Total Monthly Search Volume", tam.total_monthly_search_volume),
                ("P1 + P2 Search Volume", tam.p1_p2_search_volume),
                ("Current Capture Estimate", tam.current_capture_estimate),
                ("Opportunity Gap", tam.opportunity_gap),
                ("Current Share Ratio", tam.current_share_ratio),
                ("Methodology", tam.methodology),
            ],
        )
        workbook.save(target)

    def _write_url_architecture_map(self, response: Module0Response, target: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "URL Architecture"
        self._append_key_value_rows(
            sheet,
            [
                ("Brand Name", response.request.brand_name),
                ("Website URL", str(response.request.website_url)),
                ("Business Type", response.site_classification.business_type.value),
                ("Detected Domain", response.site_classification.detected_domain),
                ("Sitemap URL", response.site_classification.sitemap_url or ""),
            ],
        )
        start_row = 7
        headers = [
            "Hierarchy Level",
            "Page Type",
            "Current URL",
            "Proposed URL",
            "Primary Keyword",
            "Secondary Keywords",
            "Search Volume",
            "Current Ranking",
            "Status",
            "Priority",
            "AI Trigger Rate",
            "Fan-Out Coverage",
        ]
        self._write_headers(sheet, start_row, headers)
        for item in response.url_architecture_map:
            sheet.append(
                [
                    item.hierarchy_level,
                    item.page_type,
                    item.current_url or "",
                    item.proposed_url,
                    item.primary_keyword,
                    ", ".join(item.secondary_keywords),
                    item.search_volume,
                    item.current_ranking or "",
                    item.status,
                    item.priority.value,
                    item.ai_answer_trigger_rate or "",
                    item.fan_out_coverage or "",
                ]
            )
        sheet.freeze_panes = f"A{start_row + 1}"
        sheet.auto_filter.ref = f"A{start_row}:L{sheet.max_row}"
        self._set_column_widths(sheet, [16, 16, 32, 38, 32, 40, 16, 16, 14, 12, 14, 16])
        workbook.save(target)

    def _write_minimum_effort_points(self, response: Module0Response, target: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Minimum Effort"
        self._write_headers(
            sheet,
            1,
            ["Proposed URL", "Primary Keyword", "Required Links", "Avg Competitor Difficulty", "Monthly Link Velocity", "Notes"],
        )
        for item in response.minimum_effort_points:
            sheet.append(
                [
                    item.proposed_url,
                    item.primary_keyword,
                    item.required_links,
                    item.average_competitor_difficulty,
                    item.monthly_link_velocity,
                    " | ".join(item.notes),
                ]
            )
        self._set_column_widths(sheet, [40, 32, 16, 24, 20, 70])
        workbook.save(target)

    def _write_ai_sov_baseline(self, response: Module0Response, target: Path) -> None:
        workbook = Workbook()
        overview = workbook.active
        overview.title = "AI SOV Overview"
        ai_sov = response.ai_sov_baseline
        self._append_key_value_rows(
            overview,
            [
                ("Status", ai_sov.status),
                ("Overall Score", ai_sov.overall_ai_sov_score or ai_sov.overall_score),
                ("Confidence Score", ai_sov.confidence_score),
                ("Methodology", ai_sov.methodology),
                ("Brand Visibility Summary", ai_sov.brand_visibility_summary),
                ("AI Answer Triggering Keywords", len(ai_sov.ai_answer_triggering_keywords)),
                ("Missing Visibility Keywords", len(ai_sov.missing_visibility_keywords)),
                ("Recommended Actions", len(ai_sov.recommended_geo_aeo_actions)),
            ],
        )
        engine_sheet = workbook.create_sheet("Engine Results")
        self._write_headers(
            engine_sheet,
            1,
            ["Engine", "Status", "Target Queries", "Cited Queries", "Score", "Notes"],
        )
        for result in ai_sov.engine_results:
            engine_sheet.append(
                [
                    result.engine,
                    result.status,
                    result.target_queries,
                    result.cited_queries,
                    result.score,
                    " | ".join(result.notes),
                ]
            )
        self._set_column_widths(engine_sheet, [22, 18, 16, 16, 12, 70])

        query_sheet = workbook.create_sheet("Query Results")
        self._write_headers(
            query_sheet,
            1,
            [
                "Engine",
                "Query",
                "Intent",
                "Brand Likely Cited",
                "Competitors Likely Cited",
                "Reason",
                "Citation Likelihood",
                "Confidence",
                "Content Gap",
                "Recommended Action",
                "Citation Domains",
                "Notes",
            ],
        )
        for result in ai_sov.query_results:
            query_sheet.append(
                [
                    result.engine,
                    result.query,
                    result.keyword_intent.value,
                    "Yes" if result.brand_likely_cited else "No",
                    ", ".join(result.competitors_likely_cited),
                    result.reason,
                    result.citation_likelihood_score,
                    result.confidence_score,
                    result.content_gap,
                    result.recommended_content_action,
                    ", ".join(result.citation_domains),
                    " | ".join(result.notes),
                ]
            )
        self._set_column_widths(query_sheet, [20, 30, 18, 16, 34, 56, 16, 14, 40, 40, 34, 70])

        competitor_sheet = workbook.create_sheet("Competitor Visibility")
        self._write_headers(
            competitor_sheet,
            1,
            ["Domain", "Name", "Estimated Visibility Score", "Confidence", "Summary"],
        )
        for result in ai_sov.competitor_visibility_comparison:
            competitor_sheet.append(
                [
                    result.domain,
                    result.name,
                    result.likely_visibility_score,
                    result.confidence_score,
                    result.summary,
                ]
            )
        self._set_column_widths(competitor_sheet, [32, 24, 18, 14, 72])

        citation_sheet = workbook.create_sheet("Citation Likelihood")
        self._write_headers(
            citation_sheet,
            1,
            ["Keyword", "Citation Likelihood", "Confidence", "Reason"],
        )
        for result in ai_sov.citation_likelihood_by_keyword:
            citation_sheet.append(
                [
                    result.keyword,
                    result.citation_likelihood_score,
                    result.confidence_score,
                    result.reason,
                ]
            )
        self._set_column_widths(citation_sheet, [34, 18, 14, 72])

        self._add_single_column_sheet(
            workbook,
            "Missing Keywords",
            "Keyword",
            ai_sov.missing_visibility_keywords or ["No missing visibility keywords."],
            width=50,
        )
        self._add_single_column_sheet(
            workbook,
            "AI Triggering Keywords",
            "Keyword",
            ai_sov.ai_answer_triggering_keywords or ["No AI-triggering keywords identified."],
            width=50,
        )
        self._add_single_column_sheet(
            workbook,
            "Recommended Actions",
            "Action",
            ai_sov.recommended_geo_aeo_actions or ["No recommended actions generated."],
            width=90,
        )
        workbook.save(target)

    def _write_fan_out_map(self, response: Module0Response, target: Path) -> None:
        workbook = Workbook()
        overview = workbook.active
        overview.title = "Fan-Out Overview"
        fan_out = response.fan_out_map
        self._append_key_value_rows(
            overview,
            [
                ("Methodology", fan_out.methodology),
                ("Average Coverage", fan_out.average_coverage),
                ("Keyword Maps", len(fan_out.keyword_maps)),
            ],
        )
        root_sheet = workbook.create_sheet("Keyword Maps")
        self._write_headers(
            root_sheet,
            1,
            ["Root Keyword", "Coverage Score", "Invisible Keywords"],
        )
        for item in fan_out.keyword_maps:
            root_sheet.append(
                [
                    item.root_keyword,
                    item.coverage_score,
                    ", ".join(item.invisible_keywords),
                ]
            )
        self._set_column_widths(root_sheet, [34, 18, 70])

        sub_sheet = workbook.create_sheet("Sub Queries")
        self._write_headers(
            sub_sheet,
            1,
            ["Root Keyword", "Sub Query", "Content Requirement", "Has Content", "Source"],
        )
        for item in fan_out.keyword_maps:
            for sub_query in item.sub_queries:
                sub_sheet.append(
                    [
                        item.root_keyword,
                        sub_query.query,
                        sub_query.content_requirement,
                        "Yes" if sub_query.has_content else "No",
                        sub_query.source,
                    ]
                )
        self._set_column_widths(sub_sheet, [30, 34, 70, 14, 18])
        workbook.save(target)

    def _write_entity_authority_baseline(self, response: Module0Response, target: Path) -> None:
        workbook = Workbook()
        overview = workbook.active
        overview.title = "Entity Baseline"
        entity = response.entity_authority_baseline
        self._append_key_value_rows(
            overview,
            [
                ("Score", entity.score),
                ("Knowledge Panel Status", entity.knowledge_panel_status),
                ("Methodology", entity.methodology),
                ("sameAs Links", len(entity.same_as_links)),
                ("Brand Mentions", len(entity.brand_mentions)),
            ],
        )
        self._add_single_column_sheet(
            workbook,
            "sameAs Links",
            "URL",
            entity.same_as_links or ["No sameAs links detected."],
            width=100,
        )
        mentions = workbook.create_sheet("Brand Mentions")
        self._write_headers(mentions, 1, ["Source", "URL", "Mention Type", "Consistent"])
        for mention in entity.brand_mentions:
            mentions.append(
                [
                    mention.source_name,
                    mention.source_url or "",
                    mention.mention_type,
                    "Yes" if mention.consistent else "No",
                ]
            )
        self._set_column_widths(mentions, [28, 50, 18, 14])
        self._add_single_column_sheet(
            workbook,
            "Consistency Gaps",
            "Gap",
            entity.consistency_gaps or ["No consistency gaps recorded."],
            width=100,
        )
        self._add_single_column_sheet(
            workbook,
            "Opportunities",
            "Opportunity",
            entity.reinforcement_opportunities or ["No reinforcement opportunities recorded."],
            width=100,
        )
        workbook.save(target)

    def _write_warnings_errors(self, response: Module0Response, target: Path) -> None:
        workbook = Workbook()
        messages = workbook.active
        messages.title = "Warnings"
        self._write_headers(messages, 1, ["Code", "Severity", "Message", "Source"])
        for item in response.warnings_errors:
            messages.append([item.code, item.severity.value, item.message, item.source])
        self._set_column_widths(messages, [22, 14, 90, 24])
        self._add_key_value_sheet(
            workbook,
            "Run Timestamps",
            [
                ("Started At", response.run_timestamps.started_at.isoformat()),
                ("Completed At", response.run_timestamps.completed_at.isoformat()),
                (
                    "Data Fresh Until",
                    response.run_timestamps.data_fresh_until.isoformat()
                    if response.run_timestamps.data_fresh_until
                    else "",
                ),
            ],
        )
        self._add_single_column_sheet(
            workbook,
            "Next Steps",
            "Next Step",
            response.next_steps or ["No next steps."],
            width=100,
        )
        workbook.save(target)

    def _write_full_workbook(self, response: Module0Response, target: Path) -> None:
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"
        self._append_key_value_rows(
            summary,
            [
                ("Brand", response.request.brand_name),
                ("Domain", response.site_classification.detected_domain),
                ("Business Type", response.site_classification.business_type.value),
                ("CMS", response.site_classification.cms),
                ("Site Scale", response.site_classification.site_scale_tier),
                ("SEMrush Status", response.semrush.status),
                ("SEMrush Source", response.semrush.data_source),
                ("Fallback Used", "Yes" if response.semrush.fallback_used else "No"),
                ("Estimated Data", "Yes" if response.semrush.is_estimated else "No"),
                ("Total URLs", response.website_profile.url_inventory.total_urls),
                ("Keyword Universe Size", len(response.master_keyword_universe)),
                ("Quick Wins", response.quick_wins.total_count),
                ("TAM", response.tam_dataset.total_monthly_search_volume),
                ("AI SOV Overall", response.ai_sov_baseline.overall_score),
                ("Entity Authority Score", response.entity_authority_baseline.score),
            ],
        )
        self._add_key_value_sheet(
            workbook,
            "Run Timestamps",
            [
                ("Started At", response.run_timestamps.started_at.isoformat()),
                ("Completed At", response.run_timestamps.completed_at.isoformat()),
                (
                    "Data Fresh Until",
                    response.run_timestamps.data_fresh_until.isoformat()
                    if response.run_timestamps.data_fresh_until
                    else "",
                ),
            ],
        )
        self._append_workbook_sections(workbook, response)
        workbook.save(target)

    def _append_workbook_sections(self, workbook: Workbook, response: Module0Response) -> None:
        self._append_cdd_section(workbook, response)
        self._append_site_classification_section(workbook, response)
        self._append_website_profile_section(workbook, response)
        self._append_semrush_section(workbook, response)
        self._append_competitive_section(workbook, response)
        self._append_keyword_cluster_section(workbook, response)
        self._append_quick_wins_section(workbook, response)
        self._append_tam_section(workbook, response)
        self._append_url_architecture_section(workbook, response)
        self._append_minimum_effort_section(workbook, response)
        self._append_ai_sov_section(workbook, response)
        self._append_fan_out_section(workbook, response)
        self._append_entity_section(workbook, response)
        self._append_warnings_section(workbook, response)

    def _append_cdd_section(self, workbook: Workbook, response: Module0Response) -> None:
        self._add_key_value_sheet(
            workbook,
            "CDD Extraction",
            [
                ("Filename", response.cdd_file.filename),
                ("Parser Used", response.cdd_extraction.parser_used),
                ("Sections Detected", ", ".join(response.cdd_extraction.sections_detected)),
                ("Text Preview", response.cdd_extraction.text_preview),
            ],
        )

    def _append_site_classification_section(self, workbook: Workbook, response: Module0Response) -> None:
        classification = response.site_classification
        self._add_key_value_sheet(
            workbook,
            "Site Classification",
            [
                ("Detected Domain", classification.detected_domain),
                ("Business Type", classification.business_type.value),
                ("Industry Category", classification.industry_category),
                ("Geographic Target", classification.geographic_target),
                ("Language", classification.language),
                ("CMS", classification.cms),
                ("CMS Version", classification.cms_version or ""),
                ("Site Scale Tier", classification.site_scale_tier),
                ("Page Builder", classification.page_builder or ""),
                ("Sitemap URL", classification.sitemap_url or ""),
                ("Theme / Template", classification.theme_or_template or ""),
                ("Confidence Score", classification.confidence_score),
                ("Active Components", ", ".join(classification.active_components)),
                ("Notes", " | ".join(classification.notes)),
            ],
        )

    def _append_website_profile_section(self, workbook: Workbook, response: Module0Response) -> None:
        self._add_key_value_sheet(
            workbook,
            "Website Profile",
            [
                ("Homepage Status", response.website_profile.homepage_status_code or ""),
                ("Final Status", response.website_profile.final_status_code or ""),
                ("Response Time (ms)", response.website_profile.response_time_ms or ""),
                ("Redirect Count", response.website_profile.redirect_count),
                ("Homepage Title", response.website_profile.homepage_title or ""),
                ("Meta Description", response.website_profile.meta_description or ""),
                ("Canonical URL", response.website_profile.canonical_url or ""),
                ("Word Count", response.website_profile.word_count),
                ("H1 Count", response.website_profile.h1_count),
                ("Primary Schema Type", response.website_profile.primary_schema_type or ""),
                ("Mobile Friendly", response.website_profile.mobile_friendly),
                ("Indexable", response.website_profile.indexable),
                ("Robots URL", response.website_profile.robots_txt.url),
                ("Sitemap Discovered", response.website_profile.sitemap.discovered),
                ("Total URLs", response.website_profile.url_inventory.total_urls),
            ],
        )

    def _append_semrush_section(self, workbook: Workbook, response: Module0Response) -> None:
        self._add_key_value_sheet(
            workbook,
            "SEMrush",
            [
                ("Configured", response.semrush.configured),
                ("Region Database", response.semrush.region_database),
                ("Status", response.semrush.status),
                ("Data Source", response.semrush.data_source),
                ("Fallback Used", response.semrush.fallback_used),
                ("Is Estimated", response.semrush.is_estimated),
                ("Keyword Limit", response.semrush.keyword_limit or ""),
                ("Estimated Traffic", response.semrush.estimated_monthly_traffic or ""),
                ("Organic Keyword Count", response.semrush.organic_keyword_count or ""),
            ],
        )

    def _append_competitive_section(self, workbook: Workbook, response: Module0Response) -> None:
        competitors = workbook.create_sheet("Competitors")
        self._write_headers(
            competitors,
            1,
            ["Competitor", "Strength", "Confidence", "Shared Keywords", "Anomaly Filtered", "Keyword Sample"],
        )
        for item in response.competitive_intelligence.top_competitors:
            competitors.append(
                [
                    f"{item.name} ({item.domain})",
                    item.estimated_strength,
                    item.confidence_score,
                    item.shared_keywords or "",
                    "Yes" if item.anomaly_filtered else "No",
                    ", ".join(item.keyword_sample),
                ]
            )
        self._set_column_widths(competitors, [38, 14, 14, 18, 18, 60])

    def _append_keyword_cluster_section(self, workbook: Workbook, response: Module0Response) -> None:
        keyword_universe = workbook.create_sheet("Keyword Universe")
        self._write_headers(
            keyword_universe,
            1,
            ["Keyword", "Cluster ID", "Intent", "Priority", "Volume", "KD", "CPC", "Position", "Mapped URL", "Confidence", "Quality", "Estimated"],
        )
        for item in response.master_keyword_universe:
            keyword_universe.append(
                [
                    item.keyword,
                    item.cluster_id or "",
                    item.intent.value,
                    item.priority.value,
                    item.search_volume,
                    item.keyword_difficulty,
                    item.cpc or "",
                    item.current_position or "",
                    item.mapped_url or "",
                    item.confidence_score,
                    item.quality_score,
                    "Yes" if item.is_estimated else "No",
                ]
            )
        self._set_column_widths(keyword_universe, [34, 22, 18, 12, 12, 12, 10, 12, 34, 14, 14, 12])

        clusters = workbook.create_sheet("Keyword Clusters")
        self._write_headers(
            clusters,
            1,
            ["Cluster ID", "Label", "Intent", "Primary Keyword", "Keywords", "Search Volume", "Suggested URL"],
        )
        for cluster in response.keyword_clusters:
            clusters.append(
                [
                    cluster.cluster_id,
                    cluster.label,
                    cluster.intent.value,
                    cluster.primary_keyword,
                    ", ".join(cluster.keywords),
                    cluster.total_search_volume,
                    cluster.suggested_url or "",
                ]
            )
        self._set_column_widths(clusters, [24, 24, 18, 30, 60, 16, 38])

    def _append_quick_wins_section(self, workbook: Workbook, response: Module0Response) -> None:
        quick_wins = workbook.create_sheet("Quick Wins")
        self._write_headers(
            quick_wins,
            1,
            ["Keyword", "Priority", "Intent", "Volume", "KD", "Position", "Mapped URL"],
        )
        for item in response.quick_wins.keywords:
            quick_wins.append(
                [
                    item.keyword,
                    item.priority.value,
                    item.intent.value,
                    item.search_volume,
                    item.keyword_difficulty,
                    item.current_position or "",
                    item.mapped_url or "",
                ]
            )
        self._set_column_widths(quick_wins, [36, 12, 18, 12, 12, 12, 34])

    def _append_tam_section(self, workbook: Workbook, response: Module0Response) -> None:
        self._add_key_value_sheet(
            workbook,
            "TAM",
            [
                ("Total Monthly Search Volume", response.tam_dataset.total_monthly_search_volume),
                ("P1 + P2 Search Volume", response.tam_dataset.p1_p2_search_volume),
                ("Current Capture Estimate", response.tam_dataset.current_capture_estimate),
                ("Opportunity Gap", response.tam_dataset.opportunity_gap),
                ("Current Share Ratio", response.tam_dataset.current_share_ratio),
                ("Methodology", response.tam_dataset.methodology),
            ],
        )

    def _append_url_architecture_section(self, workbook: Workbook, response: Module0Response) -> None:
        sheet = workbook.create_sheet("URL Architecture")
        self._write_headers(
            sheet,
            1,
            ["Level", "Page Type", "Current URL", "Proposed URL", "Primary Keyword", "Secondary Keywords", "Volume", "Ranking", "Priority"],
        )
        for item in response.url_architecture_map:
            sheet.append(
                [
                    item.hierarchy_level,
                    item.page_type,
                    item.current_url or "",
                    item.proposed_url,
                    item.primary_keyword,
                    ", ".join(item.secondary_keywords),
                    item.search_volume,
                    item.current_ranking or "",
                    item.priority.value,
                ]
            )
        self._set_column_widths(sheet, [10, 14, 30, 38, 30, 40, 14, 14, 12])

    def _append_minimum_effort_section(self, workbook: Workbook, response: Module0Response) -> None:
        sheet = workbook.create_sheet("Minimum Effort")
        self._write_headers(sheet, 1, ["Proposed URL", "Primary Keyword", "Required Links", "Avg Difficulty", "Link Velocity", "Notes"])
        for item in response.minimum_effort_points:
            sheet.append(
                [
                    item.proposed_url,
                    item.primary_keyword,
                    item.required_links,
                    item.average_competitor_difficulty,
                    item.monthly_link_velocity,
                    " | ".join(item.notes),
                ]
            )
        self._set_column_widths(sheet, [38, 30, 14, 16, 16, 60])

    def _append_ai_sov_section(self, workbook: Workbook, response: Module0Response) -> None:
        self._add_key_value_sheet(
            workbook,
            "AI SOV",
            [
                ("Status", response.ai_sov_baseline.status),
                ("Overall Score", response.ai_sov_baseline.overall_ai_sov_score or response.ai_sov_baseline.overall_score),
                ("Confidence Score", response.ai_sov_baseline.confidence_score),
                ("Methodology", response.ai_sov_baseline.methodology),
                ("Brand Visibility Summary", response.ai_sov_baseline.brand_visibility_summary),
                ("Missing Visibility Keywords", ", ".join(response.ai_sov_baseline.missing_visibility_keywords)),
            ],
        )

    def _append_fan_out_section(self, workbook: Workbook, response: Module0Response) -> None:
        sheet = workbook.create_sheet("Fan Out")
        self._write_headers(sheet, 1, ["Root Keyword", "Coverage Score", "Invisible Keywords"])
        for item in response.fan_out_map.keyword_maps:
            sheet.append([item.root_keyword, item.coverage_score, ", ".join(item.invisible_keywords)])
        self._set_column_widths(sheet, [32, 16, 70])

    def _append_entity_section(self, workbook: Workbook, response: Module0Response) -> None:
        self._add_key_value_sheet(
            workbook,
            "Entity Baseline",
            [
                ("Score", response.entity_authority_baseline.score),
                ("Knowledge Panel Status", response.entity_authority_baseline.knowledge_panel_status),
                ("Methodology", response.entity_authority_baseline.methodology),
                ("sameAs Links", ", ".join(response.entity_authority_baseline.same_as_links)),
                ("Consistency Gaps", " | ".join(response.entity_authority_baseline.consistency_gaps)),
            ],
        )

    def _append_warnings_section(self, workbook: Workbook, response: Module0Response) -> None:
        sheet = workbook.create_sheet("Warnings")
        self._write_headers(sheet, 1, ["Code", "Severity", "Message", "Source"])
        for item in response.warnings_errors:
            sheet.append([item.code, item.severity.value, item.message, item.source])
        self._set_column_widths(sheet, [20, 14, 90, 24])

    def _add_key_value_sheet(self, workbook: Workbook, title: str, rows: list[tuple[str, object]]) -> None:
        sheet = workbook.create_sheet(title)
        self._append_key_value_rows(sheet, rows)

    def _add_single_column_sheet(
        self,
        workbook: Workbook,
        title: str,
        header: str,
        values: list[str],
        width: int = 60,
    ) -> None:
        sheet = workbook.create_sheet(title)
        self._write_headers(sheet, 1, [header])
        for value in values:
            sheet.append([value])
        self._set_column_widths(sheet, [width])

    def _append_key_value_rows(self, sheet, rows: list[tuple[str, object]]) -> None:
        for index, (label, value) in enumerate(rows, start=1):
            sheet.append([label, value])
            sheet.cell(row=index, column=1).font = self.LABEL_FONT
        self._set_column_widths(sheet, [32, 90])

    def _write_headers(self, sheet, row_index: int, headers: list[str]) -> None:
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT

    def _set_column_widths(self, sheet, widths: list[int]) -> None:
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
